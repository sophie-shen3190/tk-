#爬取所有的跨境导航网站的网址的


import os
import subprocess
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from openpyxl import Workbook

# 检查 Chrome 是否已安装
chrome_path = 'C:\Program Files\Google\Chrome\Application\chrome.exe'
if not os.path.exists(chrome_path):
    print('启动程序错误，没有正确安装Chrome浏览器')
    exit(0)

# 设置Chrome调试模式的数据存放路径
data_dir = r'D:\youbafu\selenium'
os.makedirs(data_dir, exist_ok=True)

# 通过子进程模式启动Chrome浏览器
subprocess.Popen(f'"{chrome_path}" --remote-debugging-port=9527 --user-data-dir="{data_dir}"')

options = Options()
options.add_experimental_option("debuggerAddress", "127.0.0.1:9527")

# 指定 chromedriver 的路径
chromedriver_path = "D:\chromedriver-win64\chromedriver.exe"
s = Service(chromedriver_path)

# 尝试启动浏览器
try:
    browser = webdriver.Chrome(service=s, options=options)
except Exception as e:
    print(f"启动浏览器出错：{e}")
    exit(0)

# 更改为Google搜索结果链接
browser.get(
    'https://www.google.com/search?q=%E8%B7%A8%E5%A2%83%E5%AF%BC%E8%88%AA%E7%BD%91&sca_esv=576745885&rlz=1C1GCEU_zh-CNCN1069CN1069&sxsrf=AM9HkKkiS7boUpl43jjf1XtkVbR8K_-GkA%3A1698304298781&ei=KhE6ZaWuL82l-QaL-bLQCw&oq=%E8%B7%A8%E5%A2%83%E5%AF%BC%E8%88%AA&gs_lp=Egxnd3Mtd2l6LXNlcnAiDOi3qOWig-WvvOiIqioCCAEyBRAAGIAEMgUQABiABEi5lgFQ2ARY14EBcB54AZABA5gBtgOgAdMsqgEKMS4xNi42LjMuMbgBAcgBAPgBAagCCsICChAAGEcY1gQYsAPCAgQQIxgnwgIHECMYigUYJ8ICBRAuGIAEwgIHECMY6gIYJ8ICDBAjGIoFGBMYgAQYJ8ICCxAuGIAEGMcBGNEDwgIHEAAYigUYQ8ICBxAAGAwYgATCAgcQABgNGIAEwgIFEAAYogTiAwQYACBBiAYBkAYK&sclient=gws-wiz-serp')

title_values = []
link_values = []

# 在每次滚动之后获取新加载的元素
for _ in range(10):  # 更改滚动次数为5
    browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(2)  # 给页面加载内容留一些时间

    div_index = 1
    while True:
        try:
            title_xpath = f'//*[@id="rso"]/div[{div_index}]/div/div/div[1]/div/div/span/a/h3'
            title = browser.find_element(By.XPATH, title_xpath)
            title_values.append(title.text)

            link_xpath = f'//*[@id="rso"]/div[{div_index}]/div/div/div[1]/div/div/span/a'
            link = browser.find_element(By.XPATH, link_xpath)
            link_href = link.get_attribute('href')
            link_values.append(link_href)

            div_index += 1
        except:
            break

# 创建一个新的工作簿
wb = Workbook()

# 选择活动工作表
ws = wb.active

# 将数据添加到第一列和第二列
for i, (title, link) in enumerate(zip(title_values, link_values), start=1):
    ws.cell(row=i, column=1, value=title)
    ws.cell(row=i, column=2, value=link)

# 保存工作簿
wb.save("google1.xlsx")

print("总共发现了", len(title_values), "个标题和", len(link_values), "个链接。")

# 关闭浏览器
browser.quit()
