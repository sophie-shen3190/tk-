#注意细节：打开一个浏览器之后就要关闭别的浏览器
#②小红书每天的路径都会改变的——可能要考虑做成自动的，自动获取每天新的来替换掉
#大约1分中爬取12条
import os
import subprocess
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from openpyxl import Workbook

# 检查 Chrome 是否已安装
chrome_path = 'C:\Program Files\Google\Chrome\Application\chrome.exe'
if not os.path.exists(chrome_path):
    print('启动程序错误，没有正确安装Chrome浏览器')
    exit(0)

# 设置Chrome调试模式的数据存放路径
data_dir = r'D:\youbafu\selenium'
os.makedirs(data_dir, exist_ok=True)

# 通过子进程模式启动Chrome浏览器
subprocess.Popen(f'"{chrome_path}" --remote-debugging-port=9527 --user-data-dir="{data_dir}"')

options = Options()
options.add_experimental_option("debuggerAddress", "127.0.0.1:9527")

# 指定 chromedriver 的路径
chromedriver_path = "D:\chromedriver-win64\chromedriver.exe"
s = Service(chromedriver_path)

# 尝试启动浏览器
try:
    browser = webdriver.Chrome(service=s, options=options)
except Exception as e:
    print(f"启动浏览器出错：{e}")
    exit(0)

browser.get('https://www.xiaohongshu.com/')

count_values = []
title_values = []
link_values = []

# 在每次滚动之后获取新加载的元素
for _ in range(20):  # 根据需要调整滚动次数
    browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(2)  # 给页面加载内容留一些时间

    # 获取新加载的元素
    counts = browser.find_elements(By.XPATH,
                                   '//*[@id="exploreFeeds"]/div[2]/div[2]/div/div[4]/section[*]/div/div/div/span/span[2]')
    for count in counts:
        count_values.append(count.text)

    titles = browser.find_elements(By.XPATH, '//*[@id="exploreFeeds"]/section[*]/div/div/a/span')
    for title in titles:
        title_values.append(title.text)

    links = browser.find_elements(By.XPATH, '//*[@id="exploreFeeds"]/section[*]/div/div/a')
    for link in links:
        link_href = link.get_attribute('href')
        link_values.append(link_href)
        print(link_href)



# 用于存储每个链接页面的点赞数
likes_values = []

# 循环遍历每个链接并提取点赞数
for idx, link in enumerate(link_values, 1):
    try:
        browser.get(link)
        #time.sleep(2)  # 等待页面加载完毕

        # 提取点赞数
        like_element = browser.find_element(By.XPATH,
                                            '//*[@id="noteContainer"]/div[3]/div[3]/div[1]/div[1]/span[1]/span[2]')
        likes = like_element.text
        likes_values.append(likes)

        # 立即打印提取的点赞数
        print(f"{idx}. Link: {link} - Likes: {likes}")

    except Exception as e:
        print(f"Error accessing link {link}: {e}")
        likes_values.append("N/A")  # 如果无法获取点赞数，将其标记为 N/A

# 创建一个新的工作簿
wb = Workbook()

# 选择活动工作表
ws = wb.active

# 将数据添加到第一列、第二列和第三列
for i, (count, title, link) in enumerate(zip(count_values, title_values, link_values), start=1):
    ws.cell(row=i, column=1, value=count)
    ws.cell(row=i, column=2, value=title)
    ws.cell(row=i, column=3, value=link)

# 保存工作簿
wb.save("30.xlsx")

print("总共发现了", len(count_values), "个点赞数，", len(title_values), "个标题和", len(link_values), "个链接。")

# 关闭浏览器
browser.quit()
